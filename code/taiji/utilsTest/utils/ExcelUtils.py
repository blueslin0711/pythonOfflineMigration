# -*- coding: utf-8 -*-

# 读写2003 excel
import xlrd
import xlwt
from xlutils.copy import copy
# 读写2007 excel
import openpyxl
# 文件io
import os

from ConfigUtils import Config
from CheckUtils import StrUtil, FileUtil


class Excel():
    '''数据库连接'''
    __cof_file_name = "config.txt"


    def read_excel(self, file_name, sheet_index_list = []):
        if (file_name.endswith("xlsx")): # 2007以上
            return self.__read_excel_2007(file_name, sheet_index_list)
        elif (file_name.endswith("xls")): # 2003以上
            return self.__read_excel_2003(file_name, sheet_index_list)
        else:
            raise RuntimeError("文件格式错误！只支持xls、xlsx文件的操作")

    def write_excel(self, file_name, data_list, sheet_name="sheet"):
        if (file_name.endswith("xlsx")): # 2007以上
            self.__write_excel_2007(file_name, data_list, sheet_name)
        elif (file_name.endswith("xls")):  # 2003以上
            self.__write_excel_2003(file_name, data_list, sheet_name)
        else:
            raise RuntimeError("文件格式错误！只支持xls、xlsx文件的操作")

    def __read_excel_2007(self, file_name, sheet_index_list):
        read_list = []
        if (not FileUtil().isExist(file_name)):
            return read_list
        wb = openpyxl.load_workbook(file_name)
        sheets = wb.sheetnames
        if (len(sheet_index_list) <= 0):
            sheet_index_list = range(1, len(sheets)+1) # 全部sheet读取
        for sheet_index in sheet_index_list:
            if (isinstance(sheet_index,int) and (sheet_index in range(1, len(sheets)+1))):
                sheet = wb.worksheets[sheet_index-1]
                sheet_read_list = []
                for row in sheet.rows:
                    row_list = []
                    for cell in row:
                        value = cell.value
                        if (value is None): value = ""
                        row_list.append(value)
                    sheet_read_list.append(row_list)
            else:
                print("sheet_index_list 参数不符合规范，应该表示第几个sheet")
            read_list.append(sheet_read_list)
        print("读取数据成功！")
        return read_list

    def __read_excel_2003(self, file_name, sheet_index_list):
        read_list = []
        if (not FileUtil().isExist(file_name)):
            return read_list
        workbook = xlrd.open_workbook(file_name)
        sheets = workbook.sheet_names()
        if (len(sheet_index_list) <= 0):
            sheet_index_list = range(1, len(sheets)+1) # 全部sheet读取
        for sheet_index in sheet_index_list:
            if (isinstance(sheet_index,int) and (sheet_index in range(1, len(sheets)+1))):
                worksheet = workbook.sheet_by_name(sheets[sheet_index-1])
                sheet_read_list = []
                for i in range(0, worksheet.nrows):
                    row = worksheet.row(i)
                    row_list = []
                    for j in range(0, worksheet.ncols):
                        row_list.append(worksheet.cell_value(i, j))
                    sheet_read_list.append(row_list)
            else:
                print("sheet_index_list 参数不符合规范，应该表示第几个sheet")
            read_list.append(sheet_read_list)
        print("读取数据成功！")
        return read_list

    def __write_excel_2007(self, file_name, data_list, sheet_name):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        for i in range(0, len(data_list)):
            for j in range(0, len(data_list[i])):
                sheet.cell(row=i+1, column=j+1, value=str(data_list[i][j]))
        wb.save(file_name)
        print("写入数据成功！")

    def __write_excel_2003(self, file_name, data_list, sheet_name):
        wb = xlwt.Workbook()
        sheet = wb.add_sheet(sheet_name, cell_overwrite_ok=True)
        for i in range(0, len(data_list)):
            for j in range(0, len(data_list[i])):
                sheet.write(i, j, data_list[i][j])
        wb.save(file_name)
        print("写入数据成功！")

    def modify_excel(self, file_name, modify_data_list, sheet_index=1):
        if (file_name.endswith("xlsx")): # 2007以上
            self.__modify_excel_2007(file_name, sheet_index, modify_data_list)
        elif (file_name.endswith("xls")):  # 2003以上
            self.__modify_excel_2003(file_name, sheet_index, modify_data_list)
        else:
            raise RuntimeError("文件格式错误！只支持xls、xlsx文件的操作")

    def __modify_excel_2007(self, file_name, sheet_index, modify_data_list):
        # 校验入参
        if (not FileUtil().isExist(file_name)):
            return
        # 获取excel对象
        wb = openpyxl.load_workbook(file_name)
        sheets = wb.worksheets
        if not (isinstance(sheet_index,int) and (sheet_index in range(1, len(sheets)+1))):
            print("sheet_index参数不符合规范，参数应该表示第几个sheet")
            return
        sheet = sheets[sheet_index-1]
        # 写入修改内容
        for datas in modify_data_list:
            if (len(datas) is not 3): 
                print("数据结构不对，必须包含（行号，列号，数据数组）")
                continue
            r_start = datas[0]
            c_start = datas[1]
            data_list = datas[2]
            for i in range(len(data_list)):
                c_index = c_start
                for data in data_list[i]:
                    sheet.cell(row=r_start, column=c_index, value=data)
                    c_index += 1
                r_start += 1
        # 保存
        wb.save(file_name)
        print("修改数据成功！")

    def __modify_excel_2003(self, file_name, sheet_index, modify_data_list):
        # 校验入参
        if (not FileUtil().isExist(file_name)):
            return
        # 获取excel对象
        workbook = xlrd.open_workbook(file_name)
        excel = copy(workbook)
        sheets = workbook.sheet_names()
        if not (isinstance(sheet_index,int) and (sheet_index in range(1, len(sheets)+1))):
            print("sheet_index参数不符合规范，参数应该表示第几个sheet")
            return
        sheet = excel.get_sheet(sheet_index-1)
        # 写入修改内容
        for datas in modify_data_list:
            if (len(datas) is not 3): 
                print("数据结构不对，必须包含（行号，列号，数据数组）")
                continue
            r_start = datas[0]
            c_start = datas[1]
            data_list = datas[2]
            for i in range(len(data_list)):
                c_index = c_start
                for data in data_list[i]:
                    sheet.write(r_start-1, c_index-1, data)
                    c_index += 1
                r_start += 1
        # 保存
        excel.save(file_name)
        print("修改数据成功！")


def excel_text():
    excel = Excel()
    try:
        # 测试写入excel
        # data_list = [["名称", "价格", "出版社", "语言"],
        #          ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
        #          ["暗时间", "32.4", "人民邮电出版社", "中文"],
        #          ["暗时间", "32.4", "人民邮电出版社", "中文"],
        #          ["暗时间", "32.4", "人民邮电出版社", "中文"],
        #          ["暗时间", "32.4", "人民邮电出版社", "中文"],
        #          ["暗时间", "32.4", "人民邮电出版社", "中文"],
        #          ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
        #
        # excel.write_excel("../data/111.xlsx", data_list, sheet_name="121")
        # print(excel.read_excel("../data/111.xlsx"))
        
        # 测试修改excel
        # modify_data_list 格式：[(插入点行号,插入点列号, 数据数组->[[1,2,3],[2,3,4]]),()]
        modify_data_list = [(11, 1, [["如何高效读懂一本书2", "22.3", "机械工业出版社", "中文"],["暗时间", "32.4", "人民邮电出版社", "中文"]])
                ,(16, 1, [["如何高效读懂一本书3", "22.3", "机械工业出版社", "中文1"]])]
        excel.modify_excel("../data/111.xls", modify_data_list, sheet_index=1)

    except Exception as e:
        print(e)


def main():
    excel_text()
    # input("按回车退出！")


if __name__ == '__main__':
    main()

